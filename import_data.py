"""
Excel Data Import Script for Visma Attendance System

Imports from:
- Attendance.xlsx: Daily attendance records
- Salary.xlsx: Monthly salary data (master worker data)

Usage:
    python import_data.py              # Import from Excel files
    python import_data.py --summary    # Show database summary
"""

import sys
import os
from datetime import date, datetime
from openpyxl import load_workbook
from app import create_app
from models import db, Attendance, Salary


def import_from_excel():
    """Import attendance and salary data from Excel files."""
    app = create_app()

    attendance_file = 'Attendance.xlsx'
    salary_file = 'Salary.xlsx'

    if not os.path.exists(attendance_file):
        print(f"Error: {attendance_file} not found")
        return False

    if not os.path.exists(salary_file):
        print(f"Error: {salary_file} not found")
        return False

    with app.app_context():
        try:
            # Clear existing data
            print("Clearing existing data...")
            Attendance.query.delete()
            Salary.query.delete()
            db.session.commit()

            # === First, get worker info from Attendance.xlsx ===
            print(f"\n=== Reading worker info from {attendance_file} ===")
            wb_att = load_workbook(attendance_file, data_only=True)
            ws_att = wb_att.active

            workers = {}
            attendance_data = []

            for row in ws_att.iter_rows(min_row=2, values_only=True):
                if not row[0] or not row[2]:
                    continue

                worker_id = int(row[0])
                date_val = row[1]
                name = str(row[2]).strip()
                designation = str(row[3]).strip() if row[3] else None
                team = str(row[4]).strip() if row[4] else None
                status_raw = str(row[5]).strip() if row[5] else 'Absent'
                ot_hours = float(row[6]) if row[6] else 0
                project = str(row[7]).strip() if row[7] else None

                # Store worker info
                if worker_id not in workers:
                    workers[worker_id] = {
                        'name': name,
                        'designation': designation,
                        'team': team
                    }

                # Parse date
                if isinstance(date_val, datetime):
                    record_date = date_val.date()
                elif isinstance(date_val, date):
                    record_date = date_val
                else:
                    continue

                # Parse status
                status = 'A'
                if status_raw.lower().startswith('p'):
                    status = 'P'
                elif status_raw.lower().startswith('h'):
                    status = 'H'

                attendance_data.append({
                    'worker_id': worker_id,
                    'date': record_date,
                    'status': status,
                    'ot_hours': ot_hours,
                    'project': project
                })

            print(f"Found {len(workers)} workers, {len(attendance_data)} attendance records")

            # === Import Salary (must be first - it's the master table) ===
            print(f"\n=== Importing {salary_file} ===")
            wb_sal = load_workbook(salary_file, data_only=True)
            ws_sal = wb_sal.active

            salary_count = 0

            for row in ws_sal.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue

                worker_id = int(row[0])
                total_working_days = int(row[1]) if row[1] else 0
                ot_hours = float(row[2]) if row[2] else 0
                base_salary = float(row[3]) if row[3] else 0
                total_salary = float(row[4]) if row[4] else 0

                # Get worker info
                worker = workers.get(worker_id, {})
                name = worker.get('name', f'Worker {worker_id}')
                designation = worker.get('designation')
                team = worker.get('team')

                salary = Salary(
                    worker_id=worker_id,
                    name=name,
                    designation=designation,
                    team=team,
                    total_working_days=total_working_days,
                    ot_hours=ot_hours,
                    base_salary_per_day=base_salary,
                    total_salary=total_salary
                )
                db.session.add(salary)
                salary_count += 1
                print(f"  ID {worker_id}: {name} ({designation}) - Rs.{base_salary}/day - Total: Rs.{total_salary}")

            db.session.commit()
            print(f"Imported {salary_count} salary/worker records")

            # === Import Attendance ===
            print(f"\n=== Importing attendance records ===")
            attendance_count = 0

            for att in attendance_data:
                attendance = Attendance(
                    worker_id=att['worker_id'],
                    date=att['date'],
                    status=att['status'],
                    ot_hours=att['ot_hours'],
                    project=att['project']
                )
                db.session.add(attendance)
                attendance_count += 1

            db.session.commit()
            print(f"Imported {attendance_count} attendance records")

            print("\n=== Import Complete ===")
            return True

        except Exception as e:
            print(f"Error: {e}")
            import traceback
            traceback.print_exc()
            db.session.rollback()
            return False


def show_summary():
    """Show database summary."""
    app = create_app()

    with app.app_context():
        att_count = Attendance.query.count()
        salary_count = Salary.query.count()

        print(f"\n=== Database Summary ===")
        print(f"Workers (Salary records): {salary_count}")
        print(f"Attendance Records: {att_count}")

        print("\nWorkers:")
        salaries = Salary.query.order_by(Salary.worker_id).all()
        for s in salaries:
            att = Attendance.query.filter_by(worker_id=s.worker_id).count()
            print(f"  ID {s.worker_id}: {s.name} ({s.designation}) - {att} attendance records - Rs.{s.total_salary}")


if __name__ == '__main__':
    if len(sys.argv) > 1 and sys.argv[1] == '--summary':
        show_summary()
    else:
        import_from_excel()
