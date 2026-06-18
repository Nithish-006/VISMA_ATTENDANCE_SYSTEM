from flask import Blueprint, request, jsonify, send_file, current_app
from models import db, Salary, Attendance, Worker, compute_pay
from routes.attendance import ist_now
from decimal import Decimal
from datetime import datetime, date
from functools import wraps
import hmac
from sqlalchemy import func, case
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import calendar

salary_bp = Blueprint('salary', __name__)


def require_api_key(f):
    """Guard an endpoint with the shared FINANCE_API_KEY secret.

    The caller passes the key in the `X-API-Key` header (preferred) or an
    `api_key` query param. Comparison is constant-time to avoid leaking the key
    through timing. If no key is configured server-side the endpoint refuses all
    requests (503) rather than silently running open — security is opt-in.
    """
    @wraps(f)
    def wrapper(*args, **kwargs):
        expected = current_app.config.get('FINANCE_API_KEY')
        if not expected:
            return jsonify({'error': 'API access is not configured on the server'}), 503
        provided = request.headers.get('X-API-Key') or request.args.get('api_key', '')
        if not provided or not hmac.compare_digest(str(provided), str(expected)):
            return jsonify({'error': 'Invalid or missing API key'}), 401
        return f(*args, **kwargs)
    return wrapper


@salary_bp.route('/api/salary', methods=['GET'])
def get_all_salaries():
    """Get all salary records grouped by month."""
    return get_monthly_salaries()


@salary_bp.route('/api/salary/monthly', methods=['GET'])
def get_monthly_salaries():
    """Get monthly salary breakdown."""
    # Get all salary records ordered by date (newest first)
    records = Salary.query.order_by(
        Salary.year.desc(),
        Salary.month.desc(),
        Salary.name
    ).all()

    # Group by month
    months_data = {}
    for record in records:
        month_key = f"{record.year}-{record.month:02d}"
        if month_key not in months_data:
            months_data[month_key] = {
                'workers': [],
                'total_salary': Decimal('0')
            }

        # Every figure comes from the frozen snapshot, never recomputed from the
        # live pay model. base_pay is days x the rate that was in force that
        # month; the stored total already reflects whether OT was paid, so ot_pay
        # falls out as the remainder (0 for months a worker was monthly-salaried).
        # monthly_salaried is the pay model recorded for THAT month.
        base_salary = float(record.base_salary_per_day) if record.base_salary_per_day else 0
        working_days = record.total_working_days or 0
        ot_hours = float(record.ot_hours) if record.ot_hours else 0
        total_salary = float(record.total_salary) if record.total_salary else 0
        base_pay = round(working_days * base_salary, 2)
        ot_pay = max(0.0, round(total_salary - base_pay, 2))

        months_data[month_key]['workers'].append({
            'id': record.id,
            'worker_id': record.worker_id,
            'name': record.name,
            'designation': record.designation,
            'base_salary_per_day': base_salary,
            'monthly_salaried': bool(record.monthly_salaried),
            'working_days': working_days,
            'ot_hours': ot_hours,
            'base_pay': base_pay,
            'ot_pay': ot_pay,
            'total_salary': total_salary
        })
        months_data[month_key]['total_salary'] += Decimal(str(record.total_salary or 0))

    # Convert to list
    result = []
    for month_key in sorted(months_data.keys(), reverse=True):
        year, month = month_key.split('-')
        result.append({
            'month': month_key,
            'year': int(year),
            'month_num': int(month),
            'month_name': f"{calendar.month_name[int(month)]} {year}",
            'workers': months_data[month_key]['workers'],
            'total_salary': float(months_data[month_key]['total_salary'])
        })

    return jsonify(result)


@salary_bp.route('/api/salary/project', methods=['GET'])
@require_api_key
def get_project_salary():
    """Salary master feed for one project, scoped and broken down by month.

    Built for an external finance app to assemble its own master report /
    dashboard. The `salary` table is one row per worker per MONTH and carries no
    project, so it can't be filtered by project on its own; the project lives on
    `attendance`. We therefore scope attendance to the project, then price each
    present-day with that month's stored base pay/day from the salary table
    (falling back to the worker master rate when a month has none). This is the
    same pricing the salary dashboard and the exported report use, so the figures
    line up exactly.

    Query params:
      project      - REQUIRED unless project_id given. The exact canonical
                     attendance.project value, e.g. "123 - GANTRY CRANE".
      project_id   - Alternative to project: matches every project whose value
                     starts with "{id} - " (the canonical "{id} - {name}" form).
      year, month  - Optional. Restrict to a single month (month is 1-12).
      start_date,
      end_date     - Optional YYYY-MM-DD range (ignored if year/month given).
      include_workers - "false" to omit the per-worker rows (lighter summary).

    Returns one object: project label, the resolved period, period-wide totals
    (headcount of distinct present workers, present-days, OT hours, base/OT/total
    pay) and a `months` array, each month carrying the same totals plus, by
    default, a `workers` array.
    """
    project = request.args.get('project', '').strip()
    project_id = request.args.get('project_id', '').strip()
    year_str = request.args.get('year', '').strip()
    month_str = request.args.get('month', '').strip()
    start_date_str = request.args.get('start_date', '').strip()
    end_date_str = request.args.get('end_date', '').strip()
    include_workers = request.args.get('include_workers', 'true').lower() != 'false'

    if not project and not project_id:
        return jsonify({'error': 'project (or project_id) is required'}), 400

    att_query = Attendance.query
    if project:
        att_query = att_query.filter(Attendance.project == project)
    else:
        # Match the canonical "{id} - {name}" form by its leading id segment.
        att_query = att_query.filter(Attendance.project.like(f'{project_id} - %'))

    # Optional time scoping: a single (year, month) wins; otherwise a date range.
    if year_str:
        if not year_str.isdigit():
            return jsonify({'error': 'year must be a number'}), 400
        att_query = att_query.filter(func.year(Attendance.date) == int(year_str))
        if month_str:
            if not month_str.isdigit() or not 1 <= int(month_str) <= 12:
                return jsonify({'error': 'month must be 1-12'}), 400
            att_query = att_query.filter(func.month(Attendance.date) == int(month_str))
    elif start_date_str or end_date_str:
        try:
            if start_date_str:
                att_query = att_query.filter(
                    Attendance.date >= datetime.strptime(start_date_str, '%Y-%m-%d').date())
            if end_date_str:
                att_query = att_query.filter(
                    Attendance.date <= datetime.strptime(end_date_str, '%Y-%m-%d').date())
        except ValueError:
            return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

    att_records = att_query.order_by(Attendance.date, Attendance.worker_id).all()

    # Worker identity (name / designation / rate / pay model) from the master.
    worker_ids = sorted({a.worker_id for a in att_records})
    winfo = {}
    if worker_ids:
        for w in Worker.query.filter(Worker.id.in_(worker_ids)).all():
            winfo[w.id] = {
                'name': w.name,
                'designation': w.designation or '',
                'rate': float(w.base_salary_per_day) if w.base_salary_per_day else 0,
                'monthly_salaried': bool(w.monthly_salaried),
            }

    # Per-month rate: the base pay/day in force each month is stored on that
    # month's salary row, so a worker raised mid-life keeps the old rate on
    # earlier months. Keyed by (worker_id, year, month) -> (base, monthly).
    month_rate = {}
    if worker_ids:
        for s in Salary.query.filter(Salary.worker_id.in_(worker_ids)).all():
            base = float(s.base_salary_per_day) if s.base_salary_per_day else 0
            month_rate[(s.worker_id, s.year, s.month)] = (base, bool(s.monthly_salaried))

    def basis_for(a):
        """(rate, monthly_salaried) for one record, using that month's stored rate.

        Falls back to the worker's current master rate when a month has no stored
        rate (or it's 0) — a genuinely rate-less worker stays at 0.
        """
        info = winfo.get(a.worker_id, {})
        master_rate = info.get('rate', 0)
        key = (a.worker_id, a.date.year, a.date.month)
        if key in month_rate:
            rate, monthly = month_rate[key]
            return (rate if rate > 0 else master_rate), monthly
        return master_rate, info.get('monthly_salaried', False)

    # Aggregate per month, and per worker within each month.
    months = {}
    for a in att_records:
        mkey = (a.date.year, a.date.month)
        m = months.setdefault(mkey, {'workers': {}, 'present_dates': set()})
        rate, monthly = basis_for(a)
        ot = float(a.ot_hours) if a.ot_hours else 0

        info = winfo.get(a.worker_id, {})
        w = m['workers'].setdefault(a.worker_id, {
            'worker_id': a.worker_id,
            'name': info.get('name', f'Worker {a.worker_id}'),
            'designation': info.get('designation', ''),
            'monthly_salaried': monthly,
            'rate': rate,
            'present_days': 0, 'absent_days': 0, 'holiday_days': 0,
            'ot_hours': 0.0, 'base_pay': 0.0, 'ot_pay': 0.0,
        })
        w['rate'] = rate  # that month's rate is stable for the worker
        w['monthly_salaried'] = monthly
        if a.status == 'P':
            w['present_days'] += 1
            w['base_pay'] += rate
            m['present_dates'].add(a.date)
        elif a.status == 'A':
            w['absent_days'] += 1
        elif a.status == 'H':
            w['holiday_days'] += 1
        w['ot_hours'] += ot
        # OT is paid only to daily-rate workers, at the hourly rate (rate / 8).
        if not monthly and rate > 0:
            w['ot_pay'] += (rate / 8) * ot

    # Shape the response, newest month first.
    months_out = []
    g_headcount = set()
    g_present = g_ot = g_base = g_otpay = 0
    for (yr, mo) in sorted(months.keys(), reverse=True):
        m = months[(yr, mo)]
        workers_out = []
        m_present = m_ot = m_base = m_otpay = 0
        for wid in sorted(m['workers'], key=lambda i: m['workers'][i]['name']):
            w = m['workers'][wid]
            base_pay = round(w['base_pay'], 2)
            ot_pay = round(w['ot_pay'], 2)
            workers_out.append({
                'worker_id': w['worker_id'],
                'name': w['name'],
                'designation': w['designation'],
                'monthly_salaried': w['monthly_salaried'],
                'base_salary_per_day': round(w['rate'], 2),
                'present_days': w['present_days'],
                'absent_days': w['absent_days'],
                'holiday_days': w['holiday_days'],
                'ot_hours': round(w['ot_hours'], 2),
                'base_pay': base_pay,
                'ot_pay': ot_pay,
                'total_salary': round(base_pay + ot_pay, 2),
            })
            m_present += w['present_days']
            m_ot += w['ot_hours']
            m_base += base_pay
            m_otpay += ot_pay
            g_headcount.add(wid)

        g_present += m_present
        g_ot += m_ot
        g_base += m_base
        g_otpay += m_otpay

        month_obj = {
            'year': yr,
            'month': mo,
            'month_key': f'{yr}-{mo:02d}',
            'month_name': f'{calendar.month_name[mo]} {yr}',
            'headcount': len(m['workers']),
            'working_days': len(m['present_dates']),
            'total_present_days': m_present,
            'total_ot_hours': round(m_ot, 2),
            'total_base_pay': round(m_base, 2),
            'total_ot_pay': round(m_otpay, 2),
            'total_salary': round(m_base + m_otpay, 2),
        }
        if include_workers:
            month_obj['workers'] = workers_out
        months_out.append(month_obj)

    return jsonify({
        'project': project or f'{project_id} - *',
        'project_id': project_id or (project.split(' - ', 1)[0] if ' - ' in project else None),
        'headcount': len(g_headcount),
        'total_present_days': g_present,
        'total_ot_hours': round(g_ot, 2),
        'total_base_pay': round(g_base, 2),
        'total_ot_pay': round(g_otpay, 2),
        'total_salary': round(g_base + g_otpay, 2),
        'month_count': len(months_out),
        'months': months_out,
    })


@salary_bp.route('/api/salary/<int:record_id>', methods=['GET'])
def get_salary_record(record_id):
    """Get a specific salary record."""
    salary = Salary.query.get_or_404(record_id)
    return jsonify(salary.to_dict())


@salary_bp.route('/api/salary/<int:record_id>', methods=['PUT'])
def update_salary(record_id):
    """Update salary info for a record."""
    salary = Salary.query.get_or_404(record_id)
    data = request.get_json()

    if 'base_salary_per_day' in data:
        salary.base_salary_per_day = data['base_salary_per_day']
        # Recalculate total for this month (OT excluded for monthly workers).
        worker = Worker.query.get(salary.worker_id)
        is_monthly = bool(worker.monthly_salaried) if worker else False
        _, _, salary.total_salary = compute_pay(
            salary.base_salary_per_day, salary.total_working_days, salary.ot_hours,
            monthly_salaried=is_monthly
        )

    if 'name' in data:
        salary.name = data['name']
    if 'designation' in data:
        salary.designation = data['designation']

    db.session.commit()
    return jsonify(salary.to_dict())


@salary_bp.route('/api/salary/worker/<int:worker_id>', methods=['PUT'])
def update_worker_salary(worker_id):
    """Update worker details — pay-impacting changes apply from this month on.

    The Worker master is the single source of truth and is always updated here,
    so the new rate / pay model drives every future month automatically.

    Name and designation are pure labels, so they propagate to every monthly
    snapshot for consistency. A change to base pay or the pay model (daily vs
    monthly) re-prices ONLY the current month and any later month — every month
    before this one stays frozen at the rate and total that were actually paid
    then. Finalized history is never rewritten.
    """
    data = request.get_json()
    base_salary = data.get('base_salary_per_day')
    designation = data.get('designation')
    name = data.get('name')
    monthly_salaried = data.get('monthly_salaried')

    if base_salary is None and designation is None and name is None and monthly_salaried is None:
        return jsonify({'error': 'At least one field (name, designation, base_salary_per_day, monthly_salaried) is required'}), 400

    worker = Worker.query.get(worker_id)
    if worker is None:
        return jsonify({'error': 'Worker not found'}), 404

    # 1. Update the master record (authoritative — drives future months).
    if name is not None:
        worker.name = name
    if designation is not None:
        worker.designation = designation
    if base_salary is not None:
        worker.base_salary_per_day = base_salary
    if monthly_salaried is not None:
        worker.monthly_salaried = bool(monthly_salaried)

    # The effective pay model after this update — drives whether OT is paid.
    is_monthly = bool(worker.monthly_salaried)
    pay_changed = base_salary is not None or monthly_salaried is not None

    # The boundary: the current month (IST). This month and every later month
    # are re-priced; everything earlier stays frozen.
    now = ist_now()
    current_ym = (now.year, now.month)

    # 2. Sync the monthly snapshots. Labels (name/designation) propagate to all;
    # base pay and total are rewritten only for the current month and forward.
    records = Salary.query.filter_by(worker_id=worker_id).all()
    repriced = 0
    for salary in records:
        if name is not None:
            salary.name = name
        if designation is not None:
            salary.designation = designation
        if pay_changed and (salary.year, salary.month) >= current_ym:
            if base_salary is not None:
                salary.base_salary_per_day = base_salary
            salary.monthly_salaried = is_monthly
            _, _, salary.total_salary = compute_pay(
                salary.base_salary_per_day, salary.total_working_days, salary.ot_hours,
                monthly_salaried=is_monthly
            )
            repriced += 1

    db.session.commit()

    return jsonify({
        'message': f'Updated worker {worker_id} ({repriced} month(s) re-priced)',
        'worker_id': worker_id,
        'repriced_months': repriced
    })


@salary_bp.route('/api/salary/worker/<int:worker_id>', methods=['DELETE'])
def delete_worker(worker_id):
    """Delete a worker and all their attendance and salary records."""
    salary_records = Salary.query.filter_by(worker_id=worker_id).all()
    attendance_records = Attendance.query.filter_by(worker_id=worker_id).all()
    worker = Worker.query.get(worker_id)

    if not salary_records and not attendance_records and worker is None:
        return jsonify({'error': 'Worker not found'}), 404

    salary_count = len(salary_records)
    attendance_count = len(attendance_records)

    # Delete children before the parent so the worker FK (RESTRICT) is satisfied.
    for record in attendance_records:
        db.session.delete(record)
    for record in salary_records:
        db.session.delete(record)
    db.session.flush()
    if worker is not None:
        db.session.delete(worker)

    db.session.commit()

    return jsonify({
        'message': f'Worker {worker_id} deleted',
        'deleted_salary_records': salary_count,
        'deleted_attendance_records': attendance_count
    })


# --- Report styling palette (shared across every sheet) -------------------
# A single palette keeps the workbook visually consistent and is the only place
# colours are defined, so the report's look can be retuned in one spot.
_HEADER_FILL = PatternFill('solid', fgColor='1F2937')   # slate
_HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
_TITLE_FONT = Font(bold=True, size=15, color='111827')
_SUB_FONT = Font(italic=True, size=10, color='6B7280')
_SECTION_FILL = PatternFill('solid', fgColor='374151')
_SECTION_FONT = Font(bold=True, color='FFFFFF', size=11)
_TOTAL_FILL = PatternFill('solid', fgColor='FDE68A')    # amber band for totals
_TOTAL_FONT = Font(bold=True, size=11, color='111827')
_MONEY_FILL = PatternFill('solid', fgColor='FEF9C3')    # pale yellow money column
_MONTHLY_FILL = PatternFill('solid', fgColor='DBEAFE')  # blue  = monthly-salaried
_DAILY_FILL = PatternFill('solid', fgColor='DCFCE7')    # green = daily-rate
_BAND_FILL = PatternFill('solid', fgColor='F9FAFB')     # zebra striping
_WARN_FILL = PatternFill('solid', fgColor='FECACA')     # red = needs attention (no rate set)
_WARN_FONT = Font(bold=True, color='991B1B')
_STATUS_STYLE = {  # (fill, font) per attendance status
    'P': (PatternFill('solid', fgColor='D1FAE5'), Font(bold=True, color='065F46')),
    'A': (PatternFill('solid', fgColor='FEE2E2'), Font(bold=True, color='991B1B')),
    'H': (PatternFill('solid', fgColor='FEF3C7'), Font(bold=True, color='92400E')),
}
_THIN = Side(style='thin', color='D1D5DB')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal='center', vertical='center')
_LEFT = Alignment(horizontal='left', vertical='center')
_MONEY_FMT = '#,##0'


@salary_bp.route('/api/salary/export', methods=['GET'])
def export_salary_report():
    """Build a clean, filter-scoped salary + attendance report (.xlsx).

    The workbook honours exactly the filters chosen on the salary dashboard —
    the (single-month) date range, the project, and the labour — so it never
    dumps every month the way the old report did. Pay is computed from the base
    pay/day stored for each month: salary = present days x that month's rate,
    plus OT (day rate / 8 per hour) for daily-rate workers and no OT for
    monthly-salaried workers.

    Three sheets:
      1. Worker Summary    - the complete per-worker picture (pay type, day
         rate, present / absent, OT, base pay, OT pay, total salary), with
         colour highlights so monthly-salaried vs daily-rate reads at a glance.
      2. Daily Attendance  - one row per worker per day: status, OT, project and
         work-type, with colour-coded status cells.
      3. Daily & Project   - headcount per day and per-project labour cost.
    """
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    project = request.args.get('project', '').strip()
    worker_id_str = request.args.get('worker_id', '').strip()

    # The date range scopes the entire report. The dashboard always supplies it;
    # we only fall back to all-time if it is somehow missing, so a clicked
    # download never 400s.
    try:
        start_date = (datetime.strptime(start_date_str, '%Y-%m-%d').date()
                      if start_date_str else date(2000, 1, 1))
        end_date = (datetime.strptime(end_date_str, '%Y-%m-%d').date()
                    if end_date_str else ist_now().date())
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD'}), 400

    # Attendance in range, scoped to the chosen project / labour. This is the
    # single source the whole workbook is built from.
    att_query = Attendance.query.filter(
        Attendance.date >= start_date,
        Attendance.date <= end_date,
    )
    if project:
        att_query = att_query.filter(Attendance.project == project)
    if worker_id_str.isdigit():
        att_query = att_query.filter(Attendance.worker_id == int(worker_id_str))
    att_records = att_query.order_by(Attendance.date, Attendance.worker_id).all()

    # Worker identity (name / designation / rate / pay model) from the master.
    worker_ids = sorted({a.worker_id for a in att_records})
    winfo = {}
    if worker_ids:
        for w in Worker.query.filter(Worker.id.in_(worker_ids)).all():
            winfo[w.id] = {
                'name': w.name,
                'designation': w.designation or '',
                'rate': float(w.base_salary_per_day) if w.base_salary_per_day else 0,
                'monthly_salaried': bool(w.monthly_salaried),
            }

    # Per-month rate: the base pay/day that applied in each month is stored on
    # that month's salary row, so a worker raised mid-life keeps the old rate on
    # earlier months. Keyed by (worker_id, year, month) -> (base, monthly).
    month_rate = {}
    if worker_ids:
        for s in Salary.query.filter(Salary.worker_id.in_(worker_ids)).all():
            base = float(s.base_salary_per_day) if s.base_salary_per_day else 0
            month_rate[(s.worker_id, s.year, s.month)] = (base, bool(s.monthly_salaried))

    def basis_for(a):
        """(rate, monthly_salaried) for one record — that month's stored rate.

        Look up the base pay/day saved for the record's own month and use it, so
        the salary is present days x that month's rate (+ OT for daily, none for
        monthly). If a month's stored rate is missing/0, fall back to the worker's
        current master rate (a genuinely rate-less worker like SATHISH stays 0 and
        is flagged 'Rate not set').
        """
        info = winfo.get(a.worker_id, {})
        master_rate = info.get('rate', 0)
        key = (a.worker_id, a.date.year, a.date.month)
        if key in month_rate:
            rate, monthly = month_rate[key]
            return (rate if rate > 0 else master_rate), monthly
        return master_rate, info.get('monthly_salaried', False)

    def label(wid):
        """Worker as 'NAME(ROLE)' — role rides with the name, never a stray column."""
        info = winfo.get(wid, {})
        name = info.get('name', f'Worker {wid}')
        desig = info.get('designation', '')
        return f"{name}({desig.upper()})" if desig else name

    # --- Aggregate per worker (the complete picture) ---
    workers = {}
    for a in att_records:
        rate, monthly = basis_for(a)
        ot = float(a.ot_hours) if a.ot_hours else 0
        w = workers.setdefault(a.worker_id, {
            'present': 0, 'absent': 0, 'holiday': 0, 'ot': 0.0,
            'base_pay': 0.0, 'ot_pay': 0.0, 'rates': set(), 'monthly': monthly,
        })
        w['monthly'] = monthly  # pay model is stable per worker within a period
        if rate > 0:
            w['rates'].add(rate)
        if a.status == 'P':
            w['present'] += 1
            w['base_pay'] += rate
        elif a.status == 'A':
            w['absent'] += 1
        elif a.status == 'H':
            w['holiday'] += 1
        w['ot'] += ot
        # OT is paid only to daily-rate workers, at the hourly rate (day rate / 8).
        if not monthly and rate > 0:
            w['ot_pay'] += (rate / 8) * ot

    # --- Aggregate per day and per project ---
    daily = {}
    projects = {}
    for a in att_records:
        rate, monthly = basis_for(a)
        ot = float(a.ot_hours) if a.ot_hours else 0
        ds = daily.setdefault(a.date, {'present': 0, 'absent': 0, 'holiday': 0, 'ot': 0.0})
        if a.status == 'P':
            ds['present'] += 1
        elif a.status == 'A':
            ds['absent'] += 1
        elif a.status == 'H':
            ds['holiday'] += 1
        ds['ot'] += ot
        if a.status == 'P':  # project breakdown is present-only, like the dashboard
            proj = a.project or 'Unassigned'
            ps = projects.setdefault(proj, {'wids': set(), 'days': set(),
                                            'ot': 0.0, 'cost': 0.0})
            ps['wids'].add(a.worker_id)
            ps['days'].add(a.date)
            ps['ot'] += ot
            day_ot_pay = (rate / 8) * ot if (not monthly and rate > 0) else 0
            ps['cost'] += rate + day_ot_pay

    # --- Period / filter banner shared by every sheet ---
    period_text = f"Period: {start_date.strftime('%d %b %Y')} – {end_date.strftime('%d %b %Y')}"
    filt = []
    if project:
        filt.append(f"Project: {project}")
    if worker_id_str.isdigit():
        filt.append(f"Labour: {label(int(worker_id_str))}")
    filter_text = "   |   ".join(filt) if filt else "All projects & labours"

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet

    if not att_records:
        ws = wb.create_sheet(title='No Data')
        ws['A1'] = 'No attendance found for the selected filters.'
        ws['A1'].font = _TITLE_FONT
        ws['A2'] = period_text
        ws['A3'] = filter_text
        ws['A2'].font = ws['A3'].font = _SUB_FONT
    else:
        _build_worker_summary(wb, workers, label, period_text, filter_text)
        _build_daily_attendance(wb, att_records, label, period_text, filter_text)
        _build_daily_project_summary(wb, daily, projects, period_text, filter_text)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    parts = ['salary_report', start_date.isoformat(), 'to', end_date.isoformat()]
    if project:
        parts.append(project.replace(' ', '_'))
    filename = f"{'_'.join(parts)}.xlsx"
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


def _write_banner(ws, title, period_text, filter_text, ncols):
    """Title + period + filter rows merged across the sheet; returns next row."""
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
    ws.cell(row=2, column=1, value=period_text).font = _SUB_FONT
    ws.cell(row=3, column=1, value=filter_text).font = _SUB_FONT
    last_col = get_column_letter(ncols)
    for r in (1, 2, 3):
        ws.merge_cells(f'A{r}:{last_col}{r}')
    return 5  # leave row 4 blank, data/header starts at row 5


def _style_header(ws, row_idx, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER


def _build_worker_summary(wb, workers, label, period_text, filter_text):
    """Sheet 1 — the complete per-worker picture, colour-highlighted."""
    ws = wb.create_sheet(title='Worker Summary')
    headers = ['S.No', 'Worker', 'Pay Type', 'Day Rate', 'Present', 'Absent',
               'OT Hrs', 'Base Pay', 'OT Pay', 'Total Salary']
    row = _write_banner(ws, 'WORKER SALARY SUMMARY', period_text, filter_text, len(headers))

    header_row = row
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    _style_header(ws, header_row, len(headers))
    row += 1

    money_cols = {4, 8, 9, 10}  # Day Rate, Base Pay, OT Pay, Total Salary
    tot_present = tot_absent = 0
    tot_ot = tot_base = tot_otpay = tot_total = 0.0

    for s_no, wid in enumerate(sorted(workers, key=lambda i: label(i)), start=1):
        w = workers[wid]
        base_pay = round(w['base_pay'], 2)
        ot_pay = round(w['ot_pay'], 2)
        total = round(base_pay + ot_pay, 2)
        rates = {r for r in w['rates'] if r > 0}
        # No positive rate ever seen for this worker in the period means their
        # Base Pay/Day was never entered — so days x 0 = 0. Flag it loudly rather
        # than letting a real-looking zero salary slip through unnoticed.
        rate_missing = not rates
        if rate_missing:
            rate_display = 'Rate not set'
        elif len(rates) == 1:
            rate_display = next(iter(rates))
        else:
            rate_display = 'varies'
        pay_type = 'Monthly' if w['monthly'] else 'Daily'

        values = [
            s_no, label(wid), pay_type, rate_display,
            w['present'], w['absent'], round(w['ot'], 2),
            base_pay, ot_pay, total,
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = _BORDER
            cell.alignment = _LEFT if c == 2 else _CENTER
            if c in money_cols and isinstance(v, (int, float)):
                cell.number_format = _MONEY_FMT
            if row % 2 == 0:
                cell.fill = _BAND_FILL
        # Pay-type cell: blue for monthly, green for daily-rate.
        pt = ws.cell(row=row, column=3)
        pt.fill = _MONTHLY_FILL if w['monthly'] else _DAILY_FILL
        pt.font = Font(bold=True)
        # Total salary stands out.
        tcell = ws.cell(row=row, column=10)
        tcell.fill = _MONEY_FILL
        tcell.font = Font(bold=True)
        # Missing-rate rows: flag the rate AND the (meaningless) totals in red so
        # the viewer reads "fix the rate", not "this person earned nothing".
        if rate_missing:
            for c in (4, 8, 10):
                wc = ws.cell(row=row, column=c)
                wc.fill = _WARN_FILL
                wc.font = _WARN_FONT

        tot_present += w['present']
        tot_absent += w['absent']
        tot_ot += w['ot']
        tot_base += base_pay
        tot_otpay += ot_pay
        tot_total += total
        row += 1

    # Totals band
    totals = ['', 'TOTAL', '', '', tot_present, tot_absent, round(tot_ot, 2),
              round(tot_base, 2), round(tot_otpay, 2), round(tot_total, 2)]
    for c, v in enumerate(totals, start=1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill = _TOTAL_FILL
        cell.font = _TOTAL_FONT
        cell.border = _BORDER
        cell.alignment = _LEFT if c == 2 else _CENTER
        if c in money_cols and isinstance(v, (int, float)):
            cell.number_format = _MONEY_FMT

    widths = [6, 30, 10, 11, 9, 9, 9, 12, 11, 14]
    for idx, wd in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = wd
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _build_daily_attendance(wb, att_records, label, period_text, filter_text):
    """Sheet 2 — one row per worker per day; status colour-coded."""
    ws = wb.create_sheet(title='Daily Attendance')
    headers = ['Date', 'Day', 'Worker', 'Status', 'OT Hrs', 'Project', 'Work']
    row = _write_banner(ws, 'DAILY ATTENDANCE DETAIL', period_text, filter_text, len(headers))

    header_row = row
    for c, h in enumerate(headers, start=1):
        ws.cell(row=row, column=c, value=h)
    _style_header(ws, header_row, len(headers))
    row += 1

    # Sorted by date, then worker label (att_records already date-ordered).
    ordered = sorted(att_records, key=lambda a: (a.date, label(a.worker_id)))
    for a in ordered:
        day_name = calendar.day_abbr[a.date.weekday()]
        ot = float(a.ot_hours) if a.ot_hours else 0
        values = [
            a.date.strftime('%d/%m/%Y'), day_name, label(a.worker_id),
            a.status or '', (ot if ot else ''), a.project or '-', a.work or '-',
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = _BORDER
            cell.alignment = _LEFT if c in (3, 6, 7) else _CENTER
            if row % 2 == 0:
                cell.fill = _BAND_FILL
        # Colour-code the status cell.
        scell = ws.cell(row=row, column=4)
        style = _STATUS_STYLE.get(a.status)
        if style:
            scell.fill, scell.font = style
        scell.alignment = _CENTER
        row += 1

    widths = [12, 6, 30, 8, 8, 26, 18]
    for idx, wd in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = wd
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def _build_daily_project_summary(wb, daily, projects, period_text, filter_text):
    """Sheet 3 — daily headcount and per-project labour cost."""
    ws = wb.create_sheet(title='Daily & Project')
    row = _write_banner(ws, 'DAILY & PROJECT SUMMARY', period_text, filter_text, 6)

    # --- Daily headcount section ---
    ws.cell(row=row, column=1, value='DAILY HEADCOUNT')
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = _SECTION_FILL
        ws.cell(row=row, column=c).font = _SECTION_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    day_headers = ['Date', 'Day', 'Present', 'Absent', 'Holiday', 'OT Hrs']
    for c, h in enumerate(day_headers, start=1):
        ws.cell(row=row, column=c, value=h)
    _style_header(ws, row, 6)
    row += 1

    for d in sorted(daily):
        ds = daily[d]
        values = [d.strftime('%d/%m/%Y'), calendar.day_abbr[d.weekday()],
                  ds['present'], ds['absent'], ds['holiday'], round(ds['ot'], 2)]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = _BORDER
            cell.alignment = _CENTER
            if row % 2 == 0:
                cell.fill = _BAND_FILL
        row += 1

    row += 1  # gap between sections

    # --- Project breakdown section ---
    ws.cell(row=row, column=1, value='PROJECT BREAKDOWN')
    for c in range(1, 7):
        ws.cell(row=row, column=c).fill = _SECTION_FILL
        ws.cell(row=row, column=c).font = _SECTION_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    proj_headers = ['Project', 'Workers', 'Working Days', 'OT Hrs', 'Labour Cost', '']
    for c, h in enumerate(proj_headers, start=1):
        ws.cell(row=row, column=c, value=h)
    _style_header(ws, row, 6)
    row += 1

    for proj in sorted(projects):
        ps = projects[proj]
        values = [proj, len(ps['wids']), len(ps['days']),
                  round(ps['ot'], 2), round(ps['cost'], 2), '']
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = _BORDER
            cell.alignment = _LEFT if c == 1 else _CENTER
            if c == 5 and isinstance(v, (int, float)):
                cell.number_format = _MONEY_FMT
                cell.fill = _MONEY_FILL
            elif row % 2 == 0:
                cell.fill = _BAND_FILL
        row += 1

    widths = [30, 11, 14, 10, 14, 4]
    for idx, wd in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = wd
