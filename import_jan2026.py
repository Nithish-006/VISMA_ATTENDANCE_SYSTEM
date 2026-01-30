"""
Import Jan 2026 data from Jan2026.xlsx into the database.

This script:
1. Fixes missing designations for Dec 2025 workers using Jan 2026 data
2. Updates worker names where spelling changed (HIMANSHU (B) -> (R), SURAI RAM -> SURAJ RAM)
3. Deletes existing partial Jan 2026 attendance & salary data
4. Imports full Jan 2026 attendance and salary from Excel
5. Creates new workers for those not already in the DB
"""

import sys
from datetime import date
from openpyxl import load_workbook
from app import create_app
from models import db, Attendance, Salary

# Worker mapping: Excel SNo -> DB worker_id
# Matched by name (not serial number)
WORKER_MAP = {
    1: 1,    # RAJEEV KUMAR
    2: 2,    # SURENDHAR CHOUDRY
    3: 3,    # DILIP
    4: 4,    # RAMBALAK
    5: 5,    # GORELAL
    6: 6,    # GURUDEV
    7: 7,    # NITHISH (WELDER)
    8: 8,    # KUNDAN
    9: 9,    # HIMANSHU (R) = HIMANSHU (B)
    10: 10,  # SURAJ RAM = SURAI RAM
    11: 13,  # SURAJ (FITTER)
    12: 14,  # HIMANSHU GUPTA
    13: 15,  # RAVI
    14: 16,  # DHILIP
    15: 18,  # MANUKUMAR
    16: 21,  # NITHISH (HELPER) - NEW
    17: 22,  # NITHISH KUMAR - NEW
    18: 23,  # MANISH - NEW
    19: 24,  # GAURAV - NEW
}

# Designation fixes for Dec 2025 (worker_id -> designation)
DEC_DESIGNATION_FIXES = {
    6: 'RIGGER',    # GURUDEV
    8: 'RIGGER',    # KUNDAN
    10: 'RIGGER',   # SURAI RAM / SURAJ RAM
    18: 'HELPER',   # MANUKUMAR
}

# Name updates (worker_id -> new name)
NAME_UPDATES = {
    9: 'HIMANSHU (R)',   # was HIMANSHU (B)
    10: 'SURAJ RAM',     # was SURAI RAM
}


def parse_excel():
    """Parse Jan2026.xlsx and return worker info and attendance data."""
    wb = load_workbook('Jan2026.xlsx', data_only=True)
    ws = wb['Sheet1']

    workers = {}
    attendance_records = []

    for row_idx in range(5, 24):  # rows 5-23 = workers 1-19
        sno = int(ws.cell(row=row_idx, column=1).value)
        name = str(ws.cell(row=row_idx, column=2).value).strip()
        designation = str(ws.cell(row=row_idx, column=3).value).strip() if ws.cell(row=row_idx, column=3).value else None
        worker_id = WORKER_MAP[sno]

        # Salary summary columns
        half_day_count = ws.cell(row=row_idx, column=97).value or 0
        total_present = int(ws.cell(row=row_idx, column=98).value or 0)
        total_ot = float(ws.cell(row=row_idx, column=99).value or 0)
        base_salary_per_day = float(ws.cell(row=row_idx, column=100).value or 0)
        total_salary_final = float(ws.cell(row=row_idx, column=103).value or 0)

        workers[sno] = {
            'worker_id': worker_id,
            'name': name,
            'designation': designation,
            'total_working_days': total_present,
            'ot_hours': total_ot,
            'base_salary_per_day': base_salary_per_day,
            'total_salary': total_salary_final,
        }

        # Parse daily attendance (days 1-31)
        for day in range(1, 32):
            col = 4 + (day - 1) * 3
            status_raw = ws.cell(row=row_idx, column=col).value
            ot_raw = ws.cell(row=row_idx, column=col + 1).value
            project_raw = ws.cell(row=row_idx, column=col + 2).value

            if status_raw is None:
                continue  # No data for this day

            status_str = str(status_raw).strip().upper()
            if status_str == 'P':
                status = 'P'
            elif status_str == 'H':
                status = 'H'
            else:
                status = 'A'

            ot_hours = float(ot_raw) if ot_raw else 0
            project = str(project_raw).strip() if project_raw else None

            attendance_records.append({
                'worker_id': worker_id,
                'date': date(2026, 1, day),
                'status': status,
                'ot_hours': ot_hours,
                'project': project,
            })

    return workers, attendance_records


def run_import(dry_run=False):
    """Execute the import."""
    app = create_app()

    workers, attendance_records = parse_excel()

    with app.app_context():
        try:
            # ========================================
            # STEP 1: Fix Dec 2025 designations
            # ========================================
            print("\n=== Step 1: Fix Dec 2025 missing designations ===")
            for wid, desig in DEC_DESIGNATION_FIXES.items():
                dec_records = Salary.query.filter_by(worker_id=wid, year=2025, month=12).all()
                for rec in dec_records:
                    old_desig = rec.designation
                    print(f"  Worker {wid} ({rec.name}): '{old_desig}' -> '{desig}'")
                    if not dry_run:
                        rec.designation = desig

            # ========================================
            # STEP 2: Update worker names
            # ========================================
            print("\n=== Step 2: Update worker names ===")
            for wid, new_name in NAME_UPDATES.items():
                all_records = Salary.query.filter_by(worker_id=wid).all()
                for rec in all_records:
                    old_name = rec.name
                    if old_name != new_name:
                        print(f"  Worker {wid}: '{old_name}' -> '{new_name}' (yr={rec.year}, mo={rec.month})")
                        if not dry_run:
                            rec.name = new_name

            # Also fix designations on existing Jan 2026 records that were missing
            for wid, desig in DEC_DESIGNATION_FIXES.items():
                jan_records = Salary.query.filter_by(worker_id=wid, year=2026, month=1).all()
                for rec in jan_records:
                    if not rec.designation:
                        print(f"  Worker {wid} Jan 2026 designation: '{rec.designation}' -> '{desig}'")
                        if not dry_run:
                            rec.designation = desig

            if not dry_run:
                db.session.flush()

            # ========================================
            # STEP 3: Delete existing Jan 2026 data
            # ========================================
            print("\n=== Step 3: Delete existing Jan 2026 attendance & salary ===")
            jan_att_count = Attendance.query.filter(
                Attendance.date >= date(2026, 1, 1),
                Attendance.date <= date(2026, 1, 31)
            ).count()
            jan_sal_count = Salary.query.filter_by(year=2026, month=1).count()
            print(f"  Deleting {jan_att_count} attendance records for Jan 2026")
            print(f"  Deleting {jan_sal_count} salary records for Jan 2026")

            if not dry_run:
                # Delete attendance first, then salary
                Attendance.query.filter(
                    Attendance.date >= date(2026, 1, 1),
                    Attendance.date <= date(2026, 1, 31)
                ).delete()
                db.session.flush()
                # Temporarily disable FK checks for salary delete
                # (FK on worker_id is non-unique, MySQL can't resolve partial deletes)
                db.session.execute(db.text('SET FOREIGN_KEY_CHECKS = 0'))
                Salary.query.filter_by(year=2026, month=1).delete()
                db.session.execute(db.text('SET FOREIGN_KEY_CHECKS = 1'))
                db.session.flush()

            # ========================================
            # STEP 4: Insert Jan 2026 salary records
            # ========================================
            print("\n=== Step 4: Insert Jan 2026 salary records ===")
            for sno in sorted(workers.keys()):
                w = workers[sno]
                # Get team from existing records or default to RAJU
                existing = Salary.query.filter_by(worker_id=w['worker_id']).first()
                team = existing.team if existing else 'RAJU'

                salary = Salary(
                    worker_id=w['worker_id'],
                    name=w['name'],
                    designation=w['designation'],
                    team=team,
                    base_salary_per_day=w['base_salary_per_day'],
                    year=2026,
                    month=1,
                    total_working_days=w['total_working_days'],
                    ot_hours=w['ot_hours'],
                    total_salary=w['total_salary'],
                )
                print(f"  ID {w['worker_id']:>2}: {w['name']:<22} {w['designation']:<8} "
                      f"days={w['total_working_days']:>2} ot={w['ot_hours']:>5.1f} "
                      f"base={w['base_salary_per_day']:>6.0f} total={w['total_salary']:>10.2f}")
                if not dry_run:
                    db.session.add(salary)

            if not dry_run:
                db.session.flush()

            # ========================================
            # STEP 5: Insert Jan 2026 attendance records
            # ========================================
            print(f"\n=== Step 5: Insert {len(attendance_records)} Jan 2026 attendance records ===")
            status_counts = {'P': 0, 'A': 0, 'H': 0}
            for rec in attendance_records:
                status_counts[rec['status']] += 1
                att = Attendance(
                    worker_id=rec['worker_id'],
                    date=rec['date'],
                    status=rec['status'],
                    ot_hours=rec['ot_hours'],
                    project=rec['project'],
                )
                if not dry_run:
                    db.session.add(att)

            print(f"  Present: {status_counts['P']}, Absent: {status_counts['A']}, Half-day: {status_counts['H']}")

            # ========================================
            # COMMIT
            # ========================================
            if dry_run:
                print("\n=== DRY RUN - No changes committed ===")
                db.session.rollback()
            else:
                db.session.commit()
                print("\n=== All changes committed successfully ===")

            # ========================================
            # VERIFICATION
            # ========================================
            print("\n=== Verification ===")
            jan_sal = Salary.query.filter_by(year=2026, month=1).order_by(Salary.worker_id).all()
            print(f"Jan 2026 salary records: {len(jan_sal)}")
            for s in jan_sal:
                print(f"  ID {s.worker_id:>2}: {s.name:<22} {str(s.designation):<8} "
                      f"days={s.total_working_days:>2} total={float(s.total_salary):>10.2f}")

            jan_att = Attendance.query.filter(
                Attendance.date >= date(2026, 1, 1),
                Attendance.date <= date(2026, 1, 31)
            ).count()
            print(f"Jan 2026 attendance records: {jan_att}")

            # Verify Dec 2025 designation fixes
            print("\nDec 2025 designation check:")
            for wid in DEC_DESIGNATION_FIXES:
                dec = Salary.query.filter_by(worker_id=wid, year=2025, month=12).first()
                if dec:
                    print(f"  ID {wid}: {dec.name} -> designation={dec.designation}")

            return True

        except Exception as e:
            print(f"\nERROR: {e}")
            import traceback
            traceback.print_exc()
            db.session.rollback()
            return False


if __name__ == '__main__':
    dry_run = '--dry-run' in sys.argv
    if dry_run:
        print("*** DRY RUN MODE - No changes will be saved ***")
    run_import(dry_run=dry_run)
